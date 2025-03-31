import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import json


def analyze_excel_sheet(file_path, sheet_name):
    # Load the workbook
    wb = openpyxl.load_workbook(file_path, data_only=False)
    sheet = wb[sheet_name]

    # Dictionary to store cell information
    cell_info = {}

    # Iterate through all cells
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell = sheet[f"{get_column_letter(col)}{row}"]
            cell_addr = f"{get_column_letter(col)}{row}"

            # Get cell value and formula
            value = cell.value
            formula = cell.value if str(cell.value).startswith("=") else None

            # Only store non-empty cells
            if value is not None or formula:
                cell_info[cell_addr] = {
                    "value": str(value) if value is not None else None,
                    "formula": formula,
                }

    return cell_info


def analyze_excel_with_pandas(file_path):
    # Read both sheets
    pricing_df = pd.read_excel(file_path, sheet_name="Pricing")
    blad3_df = pd.read_excel(file_path, sheet_name="Blad3")

    # Save to CSV for easier inspection
    pricing_df.to_csv("pricing_sheet.csv", index=False, encoding="utf-8")
    blad3_df.to_csv("blad3_sheet.csv", index=False, encoding="utf-8")


def main():
    file_path = "pricing.xlsx"

    # Analyze both sheets using openpyxl for formulas
    pricing_sheet = analyze_excel_sheet(file_path, "Pricing")
    blad3_sheet = analyze_excel_sheet(file_path, "Blad3")

    # Save results to JSON files for easier reading
    with open("pricing_sheet_analysis.json", "w", encoding="utf-8") as f:
        json.dump(pricing_sheet, f, indent=2, ensure_ascii=False)

    with open("blad3_sheet_analysis.json", "w", encoding="utf-8") as f:
        json.dump(blad3_sheet, f, indent=2, ensure_ascii=False)

    # Also analyze using pandas for better data structure view
    analyze_excel_with_pandas(file_path)

    print("Analysis complete. Check the following files for results:")
    print(
        "1. pricing_sheet_analysis.json and blad3_sheet_analysis.json for cell formulas"
    )
    print("2. pricing_sheet.csv and blad3_sheet.csv for data structure")


if __name__ == "__main__":
    main()
